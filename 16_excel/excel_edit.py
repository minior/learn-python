import openpyxl

wb = openpyxl.Workbook() # CREATE EMPTY WB, * Note capital W

# CELL EDITS
sheetnames = wb.sheetnames
sheet = wb['Sheet']
print(sheet['A1'].value) # None because blank workbook
sheet['A1'] = 'Hello!' # normal assign statements

print(sheet['A1'].value)

# SHEET CREATE / EDIT
sheet2 = wb.create_sheet("SheetTest")
print(sheet2.title)
sheet2.title = "Sheet2"
# Choosing Sheet Position
first_ws = wb.create_sheet("this comes first", 0) # OR index=0, title='...'
sheetnames = wb.sheetnames 
print(sheetnames)

# SAVING WB
import os
#os.chdir('change save target')
wb.save('16_excel\\excel_save_test.xlsx') # good habit to save loaded WBs in a new name (failsafe)