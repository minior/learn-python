import openpyxl
import os
#os.chdir('if needed')

workbook = openpyxl.load_workbook('16_excel\\example.xlsx')
print(workbook.sheetnames) # list of all sheets
sheet = workbook['Sheet1']
print(type(sheet))

cell = sheet['A1']
print(str(cell.value)) # returns a datetime

sheet.cell(row=1, column=2) # same result as sheet['B1'], diff use-cases
for i in range(1,8):
    print(i, sheet.cell(row=1, column=2).value) 

cell = sheet.cell(1, 2) # assumed (row, column) syntax
cell2 = sheet['B1']
print(cell.value, cell2.value) # SAME
